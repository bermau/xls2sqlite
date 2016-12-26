"""
This code is adapted from : https://gist.github.com/shahzebiam/0c2694cc17d7e39e9d12

This code uses the openpyxl package for playing around with excel using Python code
to convert complete excel workbook (all sheets) to an SQLite database
The code assumes that the first row of every sheet is the column name
Every sheet is stored in a separate table
The sheet name is assigned as the table name for every sheet

Added by BM :
- bug corrections
- tests.
- Class (intérêt pas évident ici)
v1 : classe fonctionnelle  
"""

import sqlite3
import openpyxl
from openpyxl import load_workbook
import re, os

def debug(msg, comment=''):
    print("{} : {}".format(comment, str(msg)))

def slugify(text, lower=1):
    r"""Simplify text.
    >>> slugify("  Une bé\ttise- ", lower=False)
    '_Une_bétise_'
    >>> slugify("  L'erreur potentielle- ")
    'lerreur_potentielle_'"""
    
    if lower == 1:
        text = str(text).strip().lower()
    text = re.sub(r'[^\w _-]+', '', text)
    text = re.sub(r'[- ]+', '_', text)
    return text

def unicode(x):
    return x

class WorkSheet():
    """A summary of interesting worksheets"""
    def __init__(self, name, columns):
        self.name=name
        self.header='A'
        self.columns=columns # list of column_name
    def __repr__(self):
        """Modifie la représentation d'une instance dans python"""
        return "{} instance '{}' : {}".format("WorkSheet", self.name, self.columns)
    
class Xls2Sqlite():
    """Class converting a xls file into a sqlite database file."""
    def __init__(self, xls_file, sqlite_file, new_db=True):
        pass
        self.db_name = sqlite_file
        self.xls_file=xls_file
        self.ws_dict={}
        
        if new_db and os.path.isfile(self.db_name):
            os.remove(self.db_name)
            print("{} removed.".format(self.db_name))
        #Replace with a database name
        
        self.con = sqlite3.connect(self.db_name)
        #replace with the complete path to you excel workbook
        self.wb = load_workbook(filename=self.xls_file)
        self.analyze_structure()
        self.import_structure()
        self.import_data()
        
    def analyze_structure(self):
        """Recupère les feuilles vérifie leur conformité.

met à jour le dictionnaire self.ws_dict.
self.ws_dict{'feuille1' : [1],
        'feuille3 :  [1] }"""
        sheets = self.wb.get_sheet_names()
        for sheet in sheets:
            ws = self.wb[sheet]
            if ws['A1'].value is not None:
                debug("La feuille {} est intéressante".format(sheet))
                columns = []
                for row in ws['1']:
                    columns.append(slugify(row.value))                
                ws_summary=WorkSheet(sheet, columns)
                self.ws_dict[sheet]= ws_summary
            else:
                debug("La feuille {} est sans intérêt".format(sheet))
        debug("Résumé de la structure :")

        print(self.ws_dict)
            
    def import_structure(self):
        """Import structure deducted from the first line of each sheet."""
        self.sheets = self.wb.get_sheet_names()
        for sheet in self.ws_dict:
            debug( sheet, comment="sheet")
            ws = self.wb[sheet]
            self.columns= []
            query = 'CREATE TABLE ' + slugify(sheet) + '(ID2 INTEGER PRIMARY KEY AUTOINCREMENT'
            # reading only first row
            for row in ws['1']:
                query += ', ' + slugify(row.value) + ' TEXT'
            query += ');'
            debug(query)    
            self.con.execute(query)
            self.con.commit()
                    
    def import_data(self):
        """Import data"""        
        debug("importation")
        
        # for sheet in self.ws:
        for sheet in self.ws_dict:
            debug("Traitement de : " +str(sheet))
            ws = self.wb[sheet]

            tup = []
            for i, rows in enumerate(ws):
                print(i,rows)
                tuprow = []
                if i == 0:
                    continue
                for row in rows:
                    tuprow.append(str(unicode(row.value)).strip()) if str(unicode(row.value)).strip() != 'None' else tuprow.append('')
                tup.append(tuple(tuprow))
            insQuery1 = 'INSERT INTO ' + str(slugify(sheet)) + '('
            insQuery2 = ''

            # Ici on a besoin de la liste de colonnes
            columns = self.ws_dict[sheet].columns
            for col in columns:
                insQuery1 += col + ', '
                insQuery2 += '?, '
            insQuery1 = insQuery1[:-2] + ') VALUES('
            insQuery2 = insQuery2[:-2] + ')'
            insQuery = insQuery1 + insQuery2
            debug("SQL Insert : " + insQuery)
            self.con.executemany(insQuery, tup)
            self.con.commit()
        self.con.close()
    
    def drop(self):
        pass
def demo2():

    AA = Xls2Sqlite(xls_file='essai.xls', sqlite_file='test.sqlite')

        
def _test():
    """Execute doctests."""
    import doctest
    (failures, tests) = doctest.testmod(verbose=True)
    print("{} tests performed, {} failed.".format(tests, failures))

if __name__=='__main__':

    #_test()
    demo2()
